import pandas as pd
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.styles import Font

def json_to_excel(json_file: str, excel_file: str = "job_applications.xlsx", font_size: int = 12, row_height: float = 20,
                  default_col_width: float = 25, role_col_width: float = 50):
    """
    Reads a JSON file containing a list of job applications and writes/appends them to an Excel sheet.
    Applies font size, row height, and custom column widths.
    Role column can be set wider than other columns.
    """
    
    if not Path(json_file).exists():
        print(f"JSON file '{json_file}' does not exist.")
        return
    
    # Load JSON data
    with open(json_file, "r") as f:
        applications = json.load(f)
    
    # Transform JSON to DataFrame
    data = []
    for app in applications:
        data.append({
            "Company": app.get("company", ""),
            "Role": app.get("role", ""),
            "Date Applied": app.get("date", ""),
            "Job Link/ID": app.get("job_id_or_link", "")
        })
    
    df_new = pd.DataFrame(data)
    
    # Load existing Excel or create new
    if Path(excel_file).exists():
        df_existing = pd.read_excel(excel_file)
        df_final = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_final = df_new
    
    # Save DataFrame to Excel
    df_final.to_excel(excel_file, index=False)
    
    # Load workbook with openpyxl to adjust formatting
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Set font size
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=font_size)
    
    # Set row height
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = row_height
    
    # Set column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter == 'B':  # Assuming 'Role' is column B
            ws.column_dimensions[col_letter].width = role_col_width
        else:
            ws.column_dimensions[col_letter].width = default_col_width
    
    # Save workbook
    wb.save(excel_file)
    print(f"Data successfully written to '{excel_file}' with formatting applied")
